import os
import importlib.util
import random
from flask_apscheduler import APScheduler
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify
import pandas as pd
from app_bk.ai_designer.parser import RequestParser
from app_bk.ai_designer.data_miner import DataMiner
from app_bk.ai_designer.formula_generator import FormulaGenerator
from app_bk.ai_designer.molecular_generator import MolecularGenerator
from app_bk.ai_designer.literature_miner import LiteratureMiner
from models.requirement import Requirement, RequirementDetail
from models.project import Project, ProjectStatus
from utils.auth import get_current_user, token_required
import pubchempy as pcp
from rdkit import Chem
from rdkit.Chem import Descriptors
from extensions import db
import logging

import sys
import os

# 添加项目根目录到 Python 路径
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
sys.path.insert(0, project_root)

# 现在再导入
from backend.app import injector_obj_list
from fpxh_control_sdk.plc_control_new import equipment_start

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.insert(0, project_root)
from generate_molecules.isomers.generate_isomers_from_json import main_port_generate_isomers
from ChemMind_pdfExtractor.main_port import main_port_function

old_dir = os.getcwd()
folder_path ='D:\\electrolyte\\analysis\prediction'
os.chdir(folder_path)
from analysis.prediction.prediction import MolReprAccessor, extract_feature_vectors, extract_feature_vectors_for_prediction, mlp, replace_names_with_smiles
# 导入后立即恢复目录
os.chdir(old_dir)


# 导入PostgreSQL配置 - 临时注释掉以启动项目
# from config.postgresql_config import db_manager, literature_searcher, init_database

# 简单的配方数据结构（替代数据库模型）
class SimpleFormula:
    def __init__(self, id, name, description="", system_type="",
                 application_scenario="", predicted_properties=None,
                 generation_method="", source_data=None, components=None,
                 created_at=None):
        self.id = id
        self.name = name
        self.description = description
        self.system_type = system_type
        self.application_scenario = application_scenario
        self.predicted_properties = predicted_properties or {}
        self.generation_method = generation_method
        self.source_data = source_data or {}
        self.components = components or []
        self.created_at = created_at or datetime.now().isoformat()

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'description': self.description,
            'system_type': self.system_type,
            'application_scenario': self.application_scenario,
            'predicted_properties': self.predicted_properties,
            'generation_method': self.generation_method,
            'source_data': self.source_data,
            'components': self.components,
            'created_at': self.created_at
        }

scheduler = APScheduler()    


# 创建蓝图
ai_designer_bp = Blueprint('ai_designer', __name__)

# 初始化logger并确保日志能正确输出到文件
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 明确添加文件处理器（确保日志写入文件）
try:
    # 获取日志文件路径（与app.py中的路径一致）
    log_file = os.path.join(os.path.dirname(__file__), '..', 'app.log')

    # 检查是否已经有相同的文件处理器
    has_file_handler = any(isinstance(h, logging.FileHandler) for h in logger.handlers)

    if not has_file_handler:
        # 添加文件处理器
        file_handler = logging.FileHandler(log_file, mode='a', encoding='utf-8')
        file_handler.setFormatter(logging.Formatter('[%(asctime)s] %(levelname)s: %(message)s'))
        logger.addHandler(file_handler)

        # 添加控制台处理器（方便调试）
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setFormatter(logging.Formatter('[%(asctime)s] %(levelname)s: %(message)s'))
        logger.addHandler(console_handler)

    logger.propagate = False  # 禁用传播，避免重复日志
except Exception as e:
    # 如果添加处理器失败，使用控制台输出
    print(f"警告：无法配置文件日志处理器: {e}")
    logger.addHandler(logging.StreamHandler(sys.stdout))
    logger.propagate = False

# 配方文件存储目录
FORMULA_STORAGE_DIR = os.path.join(os.path.dirname(__file__), '../data/literature/formulas')

# 初始化模块组件
request_parser = RequestParser()
data_miner = DataMiner()
formula_generator = FormulaGenerator()
molecular_generator = MolecularGenerator()
literature_miner = LiteratureMiner()

@ai_designer_bp.route('/parse-request', methods=['POST'])
def parse_user_request():
    logger.info("解析用户自然语言需求")
    try:
        data = request.get_json()
        natural_language_input = data.get("parameters", {}).get("input")

        if not natural_language_input:
            return jsonify({'error': '缺少输入文本'}), 400
        
        try:  
            # 调用deepseekAPI解析需求
            parsed_parameters = request_parser.parse_request(natural_language_input)
        except Exception as api_error:
            logger.warning(f"API解析失败")

        if not parsed_parameters:
            return jsonify({'error': '无法解析输入需求，请检查输入内容'}), 400

        return jsonify({
            'success': True,
            'data': parsed_parameters,
            'message': '需求解析成功'
        })

    except Exception as e:
        logger.error(f"解析用户需求时出错: {str(e)}")
        return jsonify({'error': f'解析失败: {str(e)}'}), 500

@ai_designer_bp.route('/confirm-parameters', methods=['POST'])
@token_required
def confirm_parameters():
    """确认并保存用户参数"""
    logger.info("确认并保存参数")
    try:
        data = request.get_json()
        logger.info(f"接收到的原始数据: {data}")

        parameters = data.get('parameters')
        logger.info(f"提取的 parameters: {parameters}, 类型: {type(parameters)}")

        if not parameters:
            logger.error("parameters 为空")
            return jsonify({'error': '缺少参数数据'}), 400

        # 检查是否为列表
        if not isinstance(parameters, list):
            logger.error(f"parameters 应该是列表类型，实际收到: {type(parameters)}")
            return jsonify({'error': '参数格式错误：应为数组格式'}), 400

        logger.info(f"接收到 {len(parameters)} 条参数记录")

        # 保存参数到数据库
        try:
            # 获取当前登录用户
            current_user = get_current_user()
            if not current_user:
                return jsonify({'error': '无法获取当前用户信息'}), 401

            # 生成 requirement_id：当前时间戳 + 自增计数
            # 先查询当前最大的 requirement_id
            max_requirement = db.session.query(db.func.max(Requirement.requirement_id)).scalar()
            max_requirement_id = max_requirement if max_requirement else 0

            # 生成新的 requirement_id（时间戳 + 自增）
            current_timestamp = int(datetime.now().timestamp() * 1000)  # 毫秒时间戳
            new_requirement_id = current_timestamp * 1000 + (max_requirement_id % 1000) + 1

            # 创建 requirement 主表记录
            requirement = Requirement(
                requirement_id=new_requirement_id,
                user_id=str(current_user.id),
                create_time=datetime.now()
            )
            db.session.add(requirement)
            db.session.flush()  # 确保获取到 id

            logger.info(f"插入 requirement 成功，requirement_id: {new_requirement_id}")

            # 定义 RequirementDetail 模型的字段白名单
            allowed_fields = {'requirement_id', 'category', 'param_name', 'param_value', 'unit', 'range', 'confidence'}

            # 遍历参数数组，创建明细记录
            detail_count = 0
            for idx, param_data in enumerate(parameters):
                try:
                    logger.info(f"处理第 {idx + 1} 条参数: {param_data}")

                    # 使用白名单过滤字段，并添加 requirement_id
                    filtered_data = {k: v for k, v in param_data.items() if k in allowed_fields}
                    filtered_data['requirement_id'] = new_requirement_id

                    # 特殊处理 confidence 字段，确保类型正确
                    if 'confidence' in filtered_data and filtered_data['confidence'] is not None:
                        try:
                            filtered_data['confidence'] = float(filtered_data['confidence'])
                        except (ValueError, TypeError):
                            filtered_data['confidence'] = 0.0

                    logger.info(f"过滤后的数据: {filtered_data}")

                    # 创建模型实例（避免使用 'range' 关键字作为构造函数参数）
                    detail = RequirementDetail()
                    for key, value in filtered_data.items():
                        setattr(detail, key, value)

                    db.session.add(detail)
                    detail_count += 1

                    logger.info(f"成功添加第 {idx + 1} 条明细记录")

                except Exception as detail_error:
                    logger.error(f"处理第 {idx + 1} 条参数时出错: {str(detail_error)}")
                    logger.error(f"错误详情: ", exc_info=True)
                    # 继续处理下一条，不中断整个流程
                    continue

            # 批量刷新所有明细记录
            if detail_count > 0:
                try:
                    db.session.flush()
                    logger.info(f"成功刷新 {detail_count} 条明细记录到数据库")
                except Exception as flush_error:
                    logger.error(f"刷新明细记录时出错: {str(flush_error)}")
                    logger.error(f"错误详情: ", exc_info=True)
                    raise flush_error  # 重新抛出异常，让外层处理

            # 初始化 project_id_str（需要在 try 外部声明）
            project_id_str = None

            # 创建 Project 记录
            try:
                # 生成 Project ID：年月日 + 10位时间戳（秒级）
                project_timestamp = int(datetime.now().timestamp())
                project_id_str = datetime.now().strftime('%Y%m%d') + str(project_timestamp)

                project = Project(
                    id=int(project_id_str),
                    name='',  # 名称为空
                    status=ProjectStatus.PENDING.value,  # 状态为0（待开始）
                    user_id=current_user.id,
                    requirement_id=requirement.requirement_id,  # 关联到 requirement 表的 requirement_id 字段
                    main_id=None,
                    created_at=datetime.now()
                )
                db.session.add(project)

                logger.info(f"创建 Project 成功，project_id: {project_id_str}, requirement_id: {requirement.requirement_id}")
            except Exception as project_error:
                logger.error(f"创建 Project 失败: {str(project_error)}")
                # Project 创建失败不影响整体流程，继续执行
                pass

            # 提交事务
            db.session.commit()
            logger.info(f"参数保存到数据库成功，共插入 {detail_count} 条明细记录")

        except Exception as db_error:
            # 回滚事务
            db.session.rollback()
            logger.error(f"保存参数到数据库失败: {str(db_error)}")
            logger.error(f"错误详情: ", exc_info=True)
            return jsonify({'error': f'保存参数失败: {str(db_error)}'}), 500

        # 参数保存成功后，立即进行文献匹配
        literature_results = []
        literature_match_success = False
        literature_error = None

        try:
            # 从前端parameters数组中提取关键词
            keywords = _extract_keywords_from_parameters(parameters)
            logger.info(f"开始文献匹配，关键词: {keywords}")

            # 调用BM25文献匹配算法
            literature_results = _call_bm25_search(keywords)

            if literature_results and len(literature_results) > 0:
                logger.info(f"BM25文献匹配完成，找到 {len(literature_results)} 篇相关文献")
                literature_match_success = True
            else:
                logger.warning("BM25文献匹配返回空结果")
                literature_results = []

        except Exception as e:
            literature_error = str(e)
            logger.error(f"BM25文献匹配失败: {literature_error}")
            logger.error(f"错误详情: ", exc_info=True)
            # 不再使用模拟数据，直接返回空结果
            literature_results = []
            logger.warning("参数保存成功，但文献匹配失败，不使用模拟数据")

        logger.info(f"确认参数接口返回: 文献数量={len(literature_results)}, 匹配成功={literature_match_success}")

        # 构建响应消息
        if literature_match_success:
            message = f'参数确认成功，文献匹配完成，找到 {len(literature_results)} 篇相关文献'
            response_data = {
                'success': True,
                'data': literature_results,
                'message': message,
                'literature_match_success': True,
                'literature_count': len(literature_results),
                'requirement_id': new_requirement_id,  # 添加 requirement_id
                'saved_parameters': parameters,  # 返回保存的参数详情
                'project_id': int(project_id_str) if 'project_id_str' in locals() else None  # 添加 project_id
            }
        else:
            message = f'参数保存成功，但文献匹配失败'
            if literature_error:
                message += f': {literature_error}'
            response_data = {
                'success': False,  # 文献匹配失败，返回False
                'data': [],
                'message': message,
                'literature_match_success': False,
                'literature_error': literature_error,
                'requirement_id': new_requirement_id,  # 添加 requirement_id
                'saved_parameters': parameters,  # 返回保存的参数详情
                'project_id': int(project_id_str) if 'project_id_str' in locals() else None  # 添加 project_id
            }

        return jsonify(response_data)

    except Exception as e:
        logger.error(f"确认参数时出错: {str(e)}")
        logger.error(f"错误详情: ", exc_info=True)
        return jsonify({'error': f'确认失败: {str(e)}'}), 500

#分子生成兜底
@ai_designer_bp.route('/mine-data1', methods=['POST'])
def underwrite_data():
    import random

    # 原始数据列表
    source_molecules = [
        {'SMILES': 'CCOC(C)=O', 'CAS': '141-78-6', 'IUPAC': 'ethyl acetate', 'common_name': 'ETHYL ACETATE', 'formula': 'C4H8O2'},
        {'SMILES': 'COCCOC', 'CAS': '110-71-4', 'IUPAC': '1,2-dimethoxyethane', 'common_name': 'Monoglyme', 'formula': 'C4H10O2'},
        {'SMILES': 'COC(C)(C)C(C)(C)C', 'CAS': '27705-21-1', 'IUPAC': '2-methoxy-2,3,3-trimethylbutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCC(C)(C)C(C)OC', 'CAS': '', 'IUPAC': '2-methoxy-3,3-dimethylpentane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCOC(C)C(C)(C)C', 'CAS': '25246-76-8', 'IUPAC': '3-ethoxy-2,2-dimethylbutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCOCC(C)(C)CC', 'CAS': '', 'IUPAC': '1-ethoxy-2,2-dimethylbutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'COCC(C)(C)COC', 'CAS': '20637-32-5', 'IUPAC': '1,3-dimethoxy-2,2-dimethylpropane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCCC(C)C(C)OC', 'CAS': '', 'IUPAC': '2-methoxy-3-methylhexane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'COC(C)C(C)C(C)C', 'CAS': '', 'IUPAC': '2-methoxy-3,4-dimethylpentane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'COCC(C)C(C)OC', 'CAS': '', 'IUPAC': '1,3-dimethoxy-2-methylbutane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCOC(C)C(C)CC', 'CAS': '', 'IUPAC': '2-ethoxy-3-methylpentane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCCOCC(C)(C)C', 'CAS': '', 'IUPAC': '2,2-dimethyl-1-propoxypropane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'COCC(F)F', 'CAS': '461-57-4', 'IUPAC': '1,1-difluoro-2-methoxyethane', 'common_name': 'difluoroethyl methyl ether', 'formula': 'C3H6F2O'},
        {'SMILES': 'COC(C)OCC(C)C', 'CAS': '', 'IUPAC': '1-(1-methoxyethoxy)-2-methylpropane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCCOC(C)C(C)C', 'CAS': '', 'IUPAC': '2-methyl-3-propoxybutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCOCCF', 'CAS': '372-96-3', 'IUPAC': '1-ethoxy-2-fluoroethane', 'common_name': None, 'formula': 'C4H9FO'},
        {'SMILES': 'FCCOCF', 'CAS': '', 'IUPAC': '1-fluoro-2-(fluoromethoxy)ethane', 'common_name': None, 'formula': 'C3H6F2O'},
        {'SMILES': 'CCOC(=O)C1CCCC1', 'CAS': '5453-85-0', 'IUPAC': 'ethyl cyclopentanecarboxylate', 'common_name': 'Ethyl cyclopentanecarboxylate', 'formula': 'C8H14O2'},
        {'SMILES': 'CCOCOC', 'CAS': '22251-34-9', 'IUPAC': 'methoxymethoxyethane', 'common_name': 'Ethoxymethoxymethane', 'formula': 'C4H10O2'},
        {'SMILES': 'CCC(=O)OC1CCCC1', 'CAS': '22499-66-7', 'IUPAC': 'cyclopentyl propanoate', 'common_name': 'Cyclopentyl propionate', 'formula': 'C8H14O2'},
        {'SMILES': 'CCOC(=O)CCC(C)C', 'CAS': '246-955-9', 'IUPAC': 'ethyl 4-methylpentanoate', 'common_name': 'Ethyl isocaproate', 'formula': 'C8H16O2'},
        {'SMILES': 'O=C(F)C(F)F', 'CAS': '2925-22-6', 'IUPAC': '2,2-difluoroacetyl fluoride', 'common_name': 'difluoroacetyl fluoride', 'formula': 'C2HF3O'},
        {'SMILES': 'CCC(C)(C)OC(C)C', 'CAS': '3249-46-5', 'IUPAC': '2-methyl-2-propan-2-yloxybutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCCC(C)(C)OCC', 'CAS': '203799-93-3', 'IUPAC': '2-ethoxy-2-methylpentane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCCOC(C)(C)CC', 'CAS': '', 'IUPAC': '2-methyl-2-propoxybutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCC(C)(C)OCOC', 'CAS': '', 'IUPAC': '2-(methoxymethoxy)-2-methylbutane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CC(C)OC(C)(C)C', 'CAS': '17348-59-3', 'IUPAC': '2-methyl-2-propan-2-yloxypropane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'COCCC(C)(C)OC', 'CAS': '39836-89-0', 'IUPAC': '1,3-dimethoxy-3-methylbutane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'COC(C)(C)CC(C)C', 'CAS': '89045-03-4', 'IUPAC': '2-methoxy-2,4-dimethylpentane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CC(=O)OC(C)(C)C', 'CAS': '540-88-5', 'IUPAC': 'tert-butyl acetate', 'common_name': 'Texaco lead appreciator', 'formula': 'C6H12O2'},
        {'SMILES': 'CCOC(C)(C)CC', 'CAS': '919-94-8', 'IUPAC': '2-ethoxy-2-methylbutane', 'common_name': 'Trimethylather', 'formula': 'C7H16O'},
        {'SMILES': 'CCCCC(C)(C)OC', 'CAS': '89045-00-1', 'IUPAC': '2-methoxy-2-methylhexane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCCOC(C)(C)C', 'CAS': '29072-93-3', 'IUPAC': '2-methyl-2-propoxypropane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'COC(C)OC(C)(C)C', 'CAS': '', 'IUPAC': '2-(1-methoxyethoxy)-2-methylpropane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'COCOC(C)(C)C', 'CAS': '24209-75-4', 'IUPAC': '2-(methoxymethoxy)-2-methylpropane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'COC(C)(C)C(C)=O', 'CAS': '36687-98-6', 'IUPAC': '3-methoxy-3-methylbutan-2-one', 'common_name': None, 'formula': 'C6H12O2'},
        {'SMILES': 'COC(C)(C)C(C)C', 'CAS': '26356-10-5', 'IUPAC': '2-methoxy-2,3-dimethylbutane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'CCC(C)(CC)OC', 'CAS': '53273-16-8', 'IUPAC': '3-methoxy-3-methylpentane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'CCCOC(=O)C(C)CC', 'CAS': '37064-20-3', 'IUPAC': 'propyl 2-methylbutanoate', 'common_name': None, 'formula': 'C8H16O2'},
        {'SMILES': 'CCCC(C)(C)OC', 'CAS': '38772-53-1', 'IUPAC': '2-methoxy-2-methylpentane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'COCC(C)(C)OC', 'CAS': '', 'IUPAC': '1,2-dimethoxy-2-methylpropane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'CCC(C)(OC)OC', 'CAS': '3453-99-4', 'IUPAC': '2,2-dimethoxybutane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'CCOCOCCOC', 'CAS': '', 'IUPAC': '1-(ethoxymethoxy)-2-methoxyethane', 'common_name': None, 'formula': 'C6H14O3'},
        {'SMILES': 'CCC(OC(C)=O)C(C)C', 'CAS': '35897-16-6', 'IUPAC': '2-methylpentan-3-yl acetate', 'common_name': None, 'formula': 'C8H16O2'},
        {'SMILES': 'CCC(C)OC(C)C', 'CAS': '18641-81-1', 'IUPAC': '2-propan-2-yloxybutane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'CCC(C)OC(C)=O', 'CAS': '105-46-4', 'IUPAC': 'butan-2-yl acetate', 'common_name': 'secbutyl acetate', 'formula': 'C6H12O2'},
        {'SMILES': 'CCC(=O)OC(C)C', 'CAS': '637-78-5', 'IUPAC': 'propan-2-yl propanoate', 'common_name': 'ISOPROPYL PROPIONATE', 'formula': 'C6H12O2'},
        {'SMILES': 'CCOC(C)C(C)C', 'CAS': '', 'IUPAC': '2-ethoxy-3-methylbutane', 'common_name': 'Isopropylather', 'formula': 'C7H16O'},
        {'SMILES': 'CCOC(CC)CC', 'CAS': '36749-13-0', 'IUPAC': '3-ethoxypentane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'COC(C)OC(C)C', 'CAS': '', 'IUPAC': '2-(1-methoxyethoxy)propane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'CCOC(C)C(C)=O', 'CAS': '1679-38-5', 'IUPAC': '3-ethoxybutan-2-one', 'common_name': 'Azetylather', 'formula': 'C6H12O2'},
        {'SMILES': 'CCCC(C)OCC', 'CAS': '1817-89-6', 'IUPAC': '2-ethoxypentane', 'common_name': 'Propylather', 'formula': 'C7H16O'},
        {'SMILES': 'CCCOC(C)CC', 'CAS': '61962-23-0', 'IUPAC': '2-propoxybutane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'CC(C)COC(C)C', 'CAS': '78448-33-6', 'IUPAC': '2-methyl-1-propan-2-yloxypropane', 'common_name': 'Isopropyl isobutyl ether', 'formula': 'C7H16O'},
        {'SMILES': 'CC(=O)COC(C)C', 'CAS': '42781-12-4', 'IUPAC': '1-propan-2-yloxypropan-2-one', 'common_name': None, 'formula': 'C6H12O2'},
        {'SMILES': 'COCCOC(C)C', 'CAS': '', 'IUPAC': '2-(2-methoxyethoxy)propane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'CCC(C)OCOC', 'CAS': '91508-78-0', 'IUPAC': '2-(methoxymethoxy)butane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'COC(=O)OC(C)C', 'CAS': '51729-83-0', 'IUPAC': 'methyl propan-2-yl carbonate', 'common_name': 'METHYL ISOPROPYL CARBONATE', 'formula': 'C5H10O3'},
        {'SMILES': 'COC(F)OC', 'CAS': '', 'IUPAC': 'fluoro(dimethoxy)methane', 'common_name': None, 'formula': 'C3H7FO2'},
        {'SMILES': 'CCOCOC(C)C', 'CAS': '', 'IUPAC': '2-(ethoxymethoxy)propane', 'common_name': 'Ethoxyisopropoxymethane', 'formula': 'C6H14O2'},
        {'SMILES': 'CCC(C)CCOC(C)=O', 'CAS': '35897-13-3', 'IUPAC': '3-methylpentyl acetate', 'common_name': None, 'formula': 'C8H16O2'},
        {'SMILES': 'COC(C)CC(C)OC', 'CAS': '41021-50-5', 'IUPAC': '2,4-dimethoxypentane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCOC(C)CCOC', 'CAS': '', 'IUPAC': '3-ethoxy-1-methoxybutane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCOCCC(C)OC', 'CAS': '', 'IUPAC': '1-ethoxy-3-methoxybutane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCOCC(C)COC', 'CAS': '', 'IUPAC': '1-ethoxy-3-methoxy-2-methylpropane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'COCC(COC)OC', 'CAS': '20637-49-4', 'IUPAC': '1,2,3-trimethoxypropane', 'common_name': 'Glycerol trimethyl ether', 'formula': 'C6H14O3'},
        {'SMILES': 'COCCCOC(C)C', 'CAS': '', 'IUPAC': '1-methoxy-3-propan-2-yloxypropane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCOCCCOCC', 'CAS': '3459-83-4', 'IUPAC': '1,3-diethoxypropane', 'common_name': 'diethoxy propane', 'formula': 'C7H16O2'},
        {'SMILES': 'CCCOCCCOC', 'CAS': '', 'IUPAC': '1-methoxy-3-propoxypropane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'COCCCOCOC', 'CAS': '', 'IUPAC': '1-methoxy-3-(methoxymethoxy)propane', 'common_name': None, 'formula': 'C6H14O3'},
        {'SMILES': 'COC(=O)OC', 'CAS': '616-38-6', 'IUPAC': 'dimethyl carbonate', 'common_name': 'Dimethyl carbonate', 'formula': 'C3H6O3'},
        {'SMILES': 'COCOCF', 'CAS': '', 'IUPAC': 'fluoromethoxy(methoxy)methane', 'common_name': None, 'formula': 'C3H7FO2'},
        {'SMILES': 'CCC(OC)C(C)C', 'CAS': '', 'IUPAC': '3-methoxy-2-methylpentane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'COC(C)C(C)(C)C', 'CAS': '', 'IUPAC': '3-methoxy-2,2-dimethylbutane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'CCC(C)C(C)OC', 'CAS': '89045-01-2', 'IUPAC': '2-methoxy-3-methylpentane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'COC(C)C(C)OC', 'CAS': '17081-24-2', 'IUPAC': '2,3-dimethoxybutane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'COC(F)(F)F', 'CAS': '421-14-7', 'IUPAC': 'trifluoro(methoxy)methane', 'common_name': 'Trifluoromethoxymethane', 'formula': 'C2H3F3O'},
        {'SMILES': 'FCOC(F)F', 'CAS': '461-63-2', 'IUPAC': 'difluoro(fluoromethoxy)methane', 'common_name': None, 'formula': 'C2H3F3O'},
        {'SMILES': 'CCC(OC)C(C)=O', 'CAS': '855381-26-9', 'IUPAC': '3-methoxypentan-2-one', 'common_name': None, 'formula': 'C6H12O2'},
        {'SMILES': 'CCC(=O)C(C)OC', 'CAS': '17042-18-1', 'IUPAC': '2-methoxypentan-3-one', 'common_name': None, 'formula': 'C6H12O2'},
        {'SMILES': 'CCCC(CC)OC', 'CAS': '54658-01-4', 'IUPAC': '3-methoxyhexane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'CCOC(C)CC(C)C', 'CAS': '', 'IUPAC': '2-ethoxy-4-methylpentane', 'common_name': 'Isobutylather', 'formula': 'C8H18O'},
        {'SMILES': 'CCOCC(C)C(C)C', 'CAS': '', 'IUPAC': '1-ethoxy-2,3-dimethylbutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'COCCC(C)COC', 'CAS': '', 'IUPAC': '1,4-dimethoxy-2-methylbutane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CC(C)CCOC(C)C', 'CAS': '', 'IUPAC': '3-methyl-1-propan-2-yloxybutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCCOCCC(C)C', 'CAS': '', 'IUPAC': '3-methyl-1-propoxybutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCC(COC)OC', 'CAS': '147818-85-7', 'IUPAC': '1,2-dimethoxybutane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'COCOCCC(C)C', 'CAS': '78400-78-9', 'IUPAC': '1-(methoxymethoxy)-3-methylbutane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCCCC(C)OCC', 'CAS': '', 'IUPAC': '2-ethoxyhexane', 'common_name': 'Butylather', 'formula': 'C8H18O'},
        {'SMILES': 'CCCC(C)COCC', 'CAS': '', 'IUPAC': '1-ethoxy-2-methylpentane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'COCCCC(C)OC', 'CAS': '', 'IUPAC': '1,4-dimethoxypentane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCCCCOCCC', 'CAS': '18641-82-2', 'IUPAC': '1-propoxypentane', 'common_name': 'Pentyl propyl ether', 'formula': 'C8H18O'},
        {'SMILES': 'CCCCCOCOC', 'CAS': '71739-39-4', 'IUPAC': '1-(methoxymethoxy)pentane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCCCCOC(C)C', 'CAS': '5756-37-6', 'IUPAC': '1-propan-2-yloxypentane', 'common_name': 'Isopropyl pentyl ether', 'formula': 'C8H18O'},
        {'SMILES': 'CCOCCCCOC', 'CAS': '', 'IUPAC': '1-ethoxy-4-methoxybutane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'COC(C)CC(C)=O', 'CAS': '13122-52-6', 'IUPAC': '4-methoxypentan-2-one', 'common_name': None, 'formula': 'C6H12O2'},
        {'SMILES': 'CCCCC(C)OC', 'CAS': '25246-71-3', 'IUPAC': '2-methoxyhexane', 'common_name': None, 'formula': 'C7H16O'},
        {'SMILES': 'COCCC(C)OC', 'CAS': '10143-66-5', 'IUPAC': '1,3-dimethoxybutane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'CCCC(C)OC(C)C', 'CAS': '', 'IUPAC': '2-propan-2-yloxypentane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCC(C)COC(C)C', 'CAS': '', 'IUPAC': '2-methyl-1-propan-2-yloxybutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCCOC(C)CCC', 'CAS': '', 'IUPAC': '2-propoxypentane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCCOCC(C)CC', 'CAS': '', 'IUPAC': '2-methyl-1-propoxybutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCCCOC(C)OC', 'CAS': '75677-94-0', 'IUPAC': '1-(1-methoxyethoxy)butane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCC(C)COCOC', 'CAS': '', 'IUPAC': '1-(methoxymethoxy)-2-methylbutane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'COC(C)C(F)OC', 'CAS': '', 'IUPAC': '1-fluoro-1,2-dimethoxypropane', 'common_name': None, 'formula': 'C5H11FO2'},
        {'SMILES': 'CCCC(C)OCOC', 'CAS': '', 'IUPAC': '2-(methoxymethoxy)pentane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'COC(=O)C(C)OC', 'CAS': '241-622-4', 'IUPAC': 'methyl 2-methoxypropanoate', 'common_name': None, 'formula': 'C5H10O3'},
        {'SMILES': 'COCC(CF)OC', 'CAS': '', 'IUPAC': '1-fluoro-2,3-dimethoxypropane', 'common_name': None, 'formula': 'C5H11FO2'},
        {'SMILES': 'CC(=O)OC1CCCC1C', 'CAS': '', 'IUPAC': '(2-methylcyclopentyl) acetate', 'common_name': None, 'formula': 'C8H14O2'},
        {'SMILES': 'CCOCC(C)OC', 'CAS': '', 'IUPAC': '1-ethoxy-2-methoxypropane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'COC(C)(OC)OC', 'CAS': '1445-45-0', 'IUPAC': '1,1,1-trimethoxyethane', 'common_name': 'Trimethyl orthoacetate', 'formula': 'C5H12O3'},
        {'SMILES': 'COC(C)OC(C)=O', 'CAS': '4382-77-8', 'IUPAC': '1-methoxyethyl acetate', 'common_name': None, 'formula': 'C5H10O3'},
        {'SMILES': 'CCOC(CC)OC', 'CAS': '127248-84-4', 'IUPAC': '1-ethoxy-1-methoxypropane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'CCCOC(C)OC', 'CAS': '', 'IUPAC': '1-(1-methoxyethoxy)propane', 'common_name': None, 'formula': 'C6H14O2'},
        {'SMILES': 'CCOC(C)OCC', 'CAS': '105-57-7', 'IUPAC': '1,1-diethoxyethane', 'common_name': 'Acetal', 'formula': 'C6H14O2'},
        {'SMILES': 'COCOC(C)OC', 'CAS': '', 'IUPAC': '1-methoxy-1-(methoxymethoxy)ethane', 'common_name': None, 'formula': 'C5H12O3'},
        {'SMILES': 'CCC(C)C(=O)OC(C)C', 'CAS': '266-411-4', 'IUPAC': 'propan-2-yl 2-methylbutanoate', 'common_name': None, 'formula': 'C8H16O2'},
        {'SMILES': 'COC(OC)(OC)OC', 'CAS': '1850-14-2', 'IUPAC': 'tetramethoxymethane', 'common_name': 'Tetramethyl orthocarbonate', 'formula': 'C5H12O4'},
        {'SMILES': 'COCOC(OC)OC', 'CAS': '', 'IUPAC': 'dimethoxy(methoxymethoxy)methane', 'common_name': None, 'formula': 'C5H12O4'},
        {'SMILES': 'COC(OC)C(C)C', 'CAS': '815-376-9', 'IUPAC': '1,1-dimethoxy-2-methylpropane', 'common_name': 'Isobutylaldehyde dimethyl acetal', 'formula': 'C6H14O2'},
        {'SMILES': 'CCOC(C)OCOC', 'CAS': '', 'IUPAC': '1-ethoxy-1-(methoxymethoxy)ethane', 'common_name': None, 'formula': 'C6H14O3'},
        {'SMILES': 'CCOC(C)(OC)OC', 'CAS': '57999-64-1', 'IUPAC': '1-ethoxy-1,1-dimethoxyethane', 'common_name': None, 'formula': 'C6H14O3'},
        {'SMILES': 'CCOCOC(C)OC', 'CAS': '', 'IUPAC': '1-(ethoxymethoxy)-1-methoxyethane', 'common_name': None, 'formula': 'C6H14O3'},
        {'SMILES': 'COC(OC)C(C)=O', 'CAS': '6342-56-9', 'IUPAC': '1,1-dimethoxypropan-2-one', 'common_name': 'Methylglyoxal dimethyl acetal', 'formula': 'C5H10O3'},
        {'SMILES': 'COCOCOCOC', 'CAS': '13353-03-2', 'IUPAC': 'methoxy(methoxymethoxymethoxy)methane', 'common_name': None, 'formula': 'C5H12O4'},
        {'SMILES': 'CCCC(OC)OC', 'CAS': '4461-87-4', 'IUPAC': '1,1-dimethoxybutane', 'common_name': 'Butanal dimethyl acetal', 'formula': 'C6H14O2'},
        {'SMILES': 'CCCOC(=O)C(C)(C)C', 'CAS': '5129-35-1', 'IUPAC': 'propyl 2,2-dimethylpropanoate', 'common_name': 'Propyl pivalate', 'formula': 'C8H16O2'},
        {'SMILES': 'CC(C)(C)OC(C)(C)C', 'CAS': '672-143-2', 'IUPAC': '2-methyl-2-[(2-methylpropan-2-yl)oxy]propane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'COC(=O)C(C)(C)C(C)C', 'CAS': '1727-57-7', 'IUPAC': 'methyl 2,2,3-trimethylbutanoate', 'common_name': None, 'formula': 'C8H16O2'},
        {'SMILES': 'CCC(C)OC(C)CC', 'CAS': '6863-58-7', 'IUPAC': '2-butan-2-yloxybutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CC(C)COC(C)(C)C', 'CAS': '33021-02-2', 'IUPAC': '2-methyl-1-[(2-methylpropan-2-yl)oxy]propane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCC(C)OCC(C)C', 'CAS': '', 'IUPAC': '2-(2-methylpropoxy)butane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCCCOC(C)(C)C', 'CAS': '1000-63-1', 'IUPAC': '1-[(2-methylpropan-2-yl)oxy]butane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CC(=O)OCCCC(C)C', 'CAS': '', 'IUPAC': '4-methylpentyl acetate', 'common_name': 'isohexyl acetate', 'formula': 'C8H16O2'},
        {'SMILES': 'CCC(C)OC(C)(C)C', 'CAS': '32970-45-9', 'IUPAC': '2-[(2-methylpropan-2-yl)oxy]butane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'CCCCOC(C)CC', 'CAS': '999-65-5', 'IUPAC': '1-butan-2-yloxybutane', 'common_name': None, 'formula': 'C8H18O'},
        {'SMILES': 'COCC(OC)OC', 'CAS': '24332-20-5', 'IUPAC': '1,1,2-trimethoxyethane', 'common_name': 'Methoxyacetaldehyde dimethyl acetal', 'formula': 'C5H12O3'},
        {'SMILES': 'CC(C)COCC(C)C', 'CAS': '628-55-7', 'IUPAC': '2-methyl-1-(2-methylpropoxy)propane', 'common_name': 'DIISOBUTYL ETHER', 'formula': 'C8H18O'},
        {'SMILES': 'CCCCOCCCC', 'CAS': '142-96-1', 'IUPAC': '1-butoxybutane', 'common_name': 'Butyl ether', 'formula': 'C8H18O'},
        {'SMILES': 'COC(OC)OC(C)C', 'CAS': '', 'IUPAC': '2-(dimethoxymethoxy)propane', 'common_name': None, 'formula': 'C6H14O3'},
        {'SMILES': 'CCOCOC(C)(C)C', 'CAS': '', 'IUPAC': '2-(ethoxymethoxy)-2-methylpropane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCOCOC(C)CC', 'CAS': '69537-82-2', 'IUPAC': '2-(ethoxymethoxy)butane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CCOC(C)OC(C)C', 'CAS': '25334-93-4', 'IUPAC': '2-(1-ethoxyethoxy)propane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'CC(C)OCOC(C)C', 'CAS': '2568-89-0', 'IUPAC': '2-(propan-2-yloxymethoxy)propane', 'common_name': 'Formaldehyde diisopropyl acetal', 'formula': 'C7H16O2'},
        {'SMILES': 'CCCOCOC(C)C', 'CAS': '', 'IUPAC': '1-(propan-2-yloxymethoxy)propane', 'common_name': None, 'formula': 'C7H16O2'},
        {'SMILES': 'COCOCOC(C)C', 'CAS': '', 'IUPAC': '2-(methoxymethoxymethoxy)propane', 'common_name': None, 'formula': 'C6H14O3'},
        {'SMILES': 'CCCOC(C)OCC', 'CAS': '20680-10-8', 'IUPAC': '1-(1-ethoxyethoxy)propane', 'common_name': 'ETHYL PROPYL ACETAL', 'formula': 'C7H16O2'},
        {'SMILES': 'CCCCOCOCC', 'CAS': '', 'IUPAC': '1-(ethoxymethoxy)butane', 'common_name': 'ethoxymethoxybutane', 'formula': 'C7H16O2'},
        {'SMILES': 'CCOC(F)(F)CC', 'CAS': '', 'IUPAC': '1-ethoxy-1,1-difluoropropane', 'common_name': None, 'formula': 'C5H10F2O'}
    ]

    def get_random_molecules(data_list, count=20):
        """
        从给定的列表中随机抽取指定数量的元素。
        如果列表长度小于 count，则返回整个列表。
        """
        num_to_select = min(count, len(data_list))
        return random.sample(data_list, num_to_select)

    # 执行函数并赋值
    molecule_data = get_random_molecules(source_molecules)
    return jsonify({
            'success': True,
            'molecules': molecule_data,
            'message': '数据挖掘完成'
        })

@ai_designer_bp.route('/mine-data', methods=['POST'])
def mine_and_augment_data():
    """文献挖掘与分子生成"""
    logger.info("文献挖掘与分子生成")

    data = request.get_json()
    literature_results = data.get('literature_results', [])
    if not literature_results:
        return jsonify({'error': '缺少文献数据'}), 400
    
    confirmed_parameters = data.get('parameters', [])

    titles = []
    #限制三篇文献进行挖掘
    i = 1
    for literature_result in literature_results:
        if i < 4:
            i = i + 1
            titles.append(literature_result.get("title"))
        else:
            break

    #step1:文献挖掘
    api_key = "sk-23734b4fae6c43febaea63d21ad747a2"
    model = "deepseek-chat"

    try:
        long_text = _build_user_request_description(confirmed_parameters)
        logger.info(f"用户需求：{long_text}")
    except Exception as e:
        logger.error(f"需求生成时出错: {str(e)}")
        return jsonify({'error': f'需求生成失败: {str(e)}'}), 500
    try:
        # user_requests = smart_truncate(long_text, 10000)
        user_requests = long_text
        input_molecules_json = main_port_function(user_requests, api_key, model, titles)
    except Exception as e:
        logger.error(f"文献挖掘出错: {str(e)}")
        return jsonify({'error': f'文献挖掘失败: {str(e)}'}), 500

    logger.info(f"文献挖掘返回的input_molecules_json类型: {type(input_molecules_json)}")
    logger.info(f"文献挖掘返回的内容: {input_molecules_json}")

    #step 2:分子生成算法
    try:
        # 解析JSON字符串为字典（因为main_port_function返回的是JSON字符串）
        import json
        if isinstance(input_molecules_json, str):
            input_molecules_dict = json.loads(input_molecules_json)
        else:
            input_molecules_dict = input_molecules_json

        #删掉锂盐部分 不用于分子生成
        input_molecules_dict["salts"] = {}
        logger.info(f"传递给分子生成函数的数据: {input_molecules_dict}")

        #调用分子生成算法
        molecules = main_port_generate_isomers(input_molecules_dict)
        logger.info(molecules)
    except Exception as e:
        logger.error(f"分子生成出错: {str(e)}")
        logger.error(f"错误详情: ", exc_info=True)
        return jsonify({'error': f'分子生成失败: {str(e)}'}), 500
    
    #先去重
    smiles_list = molecules["additives"]
    smiles_list = list(set(smiles_list))

    #获取CAS等信息
    results = []
    for idx, smiles in enumerate(smiles_list):
        try:
            lookup_result = quick_lookup(smiles)
            if lookup_result is None:
                # quick_lookup返回None时使用默认结构
                logger.warning(f"SMILES {smiles} 查询失败，使用默认结构")
                lookup_result = {
                    'SMILES': smiles,
                    'CAS': [],
                    'IUPAC': None,
                    'common_name': None,
                    'synonyms': [],
                    'formula': None,
                    'weight': None,
                    'query_error': True
                }
            results.append(lookup_result)
            logger.info(f"成功处理分子 {idx+1}/20: {smiles}")
        except Exception as e:
            logger.error(f"查询SMILES {smiles} 时出错: {str(e)}")
            # 添加错误记录
            results.append({
                'SMILES': smiles,
                'CAS': [],
                'IUPAC': None,
                'common_name': None,
                'synonyms': [],
                'formula': None,
                'weight': None,
                'query_error': True,
                'error_message': str(e)
            })

    logger.info(f"共处理 {len(results)} 个分子，其中 {sum(1 for r in results if r.get('query_error'))} 个查询失败")

    #只保留20个有查询结果的分子，不足20个算最大
    results = [result for result in results if result.get('weight')]
    results = random.sample(results, min(20,len(results)))
    logger.info(f"最终{len(results)}个分子：{results}")

    return jsonify({
            'success': True,
            'molecules': results,
            'message': f'分子生成完成，成功获取 {len(results)} 个分子的信息'
        })

@ai_designer_bp.route('/predict-properties', methods=['POST'])
def predict_properties():
    """性质预测 - 使用UniMol算法预测分子性质"""

    logger.info("开始性质预测")
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': '请提供SMILES列表'}), 400

        smiles_list = data.get('smiles_list', [])

        if not smiles_list:
            return jsonify({'error': '缺少SMILES列表'}), 400

        if not isinstance(smiles_list, list):
            return jsonify({'error': 'SMILES列表格式错误，应为数组格式'}), 400

        logger.info(f"接收到 {len(smiles_list)} 个SMILES分子待预测")
        logger.debug(f"SMILES列表前3个: {smiles_list[:3]}")

        #1. 生成csv文件，放到指定目录下
        import sys
        import os

        # 添加路径
        electrolyte_dir = r'D:\electrolyte'
        unimol_dir = os.path.join(electrolyte_dir, "unimol")
        sys.path.insert(0, unimol_dir)

        try:
            logger.info("开始导入unimol.tocvs模块")
            from tocvs import tocvs
            logger.info("unimol.tocvs模块导入成功")
        except ImportError as e:
            logger.error(f"导入unimol.tocvs失败: {e}")
            logger.error(f"sys.path: {sys.path}")
            raise

        import datetime

        # 使用方式
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        logger.info(f"生成时间戳: {timestamp}")

        logger.info("步骤1: 开始调用tocvs生成CSV文件")
        try:
            tocvs(smiles_list, timestamp)
            logger.info(f"tocvs执行完成，生成的文件路径前缀: unimol_{timestamp}")
        except Exception as e:
            logger.error(f"tocvs执行失败: {e}", exc_info=True)
            raise

        #2. 模型预测分子性质
        try:
            logger.info("开始导入unimol.run1/run2/run3模块")
            from run1 import main1
            from run2 import main2
            from run3 import main3
            logger.info("unimol.run1/run2/run3模块导入成功")
        except ImportError as e:
            logger.error(f"导入unimol.run模块失败: {e}")
            logger.error(f"sys.path: {sys.path}")
            raise

        path = f"unimol_{timestamp}"
        logger.info(f"步骤2: 开始模型预测，数据集路径: {path}")

        try:
            logger.info("步骤2.1: 开始执行main1 - 数据预处理(LMDB转换)")
            main1(path)
            logger.info("步骤2.1: main1执行完成")
        except Exception as e:
            logger.error(f"main1执行失败: {e}", exc_info=True)
            raise

        try:
            logger.info("步骤2.2: 开始执行main2 - 模型推理")
            main2(path)
            logger.info("步骤2.2: main2执行完成")
        except Exception as e:
            logger.error(f"main2执行失败: {e}", exc_info=True)
            raise

        try:
            logger.info("步骤2.3: 开始执行main3 - 结果后处理")
            main3(path)
            logger.info("步骤2.3: main3执行完成")
        except Exception as e:
            logger.error(f"main3执行失败: {e}", exc_info=True)
            raise

        logger.info(f"所有预测步骤完成")

        #3. 转为结果返回给前端
        csv_file = os.path.join(r'D:\electrolyte\unimol\final', path, f'{path}_infer_yn.csv')
        
        if not os.path.exists(csv_file):
            logger.info("文件不存在")
            return jsonify(f'error：文件{csv_file}不存在'), 500
        try:
            result = get_csv_responsen(csv_file)
        except Exception as e:
            logger.error(f"转换出错")
            return jsonify('error：转换出错'), 500 
        else:
            logger.info(result)
            return jsonify({"success": True, "result_properties": result})

    except Exception as e:
        logger.error(f"性质预测时出错: {str(e)}")
        logger.error(f"错误详情: ", exc_info=True)
        return jsonify({'error': f'预测失败: {str(e)}'}), 500 

#默认多少个配方
@ai_designer_bp.route('/generate-formula', methods=['POST'])
def generate_initial_recommendation():
    """生成初始配方"""
    logger.info("生成初始配方、配方预测")
    try:
        data = request.get_json()
        logger.info(data)
        recipe_list = data.get('recipe_list', [])
        if not recipe_list:
            return jsonify({'error': '缺少溶剂和盐数据'}), 400
        
        additive_list = data.get('additive_list', [])
        electrolyte_quantity = data.get('electrolyte_quantity', 8) 
        additive_range = data.get('additive_range')
        quality = data.get('quality')
        
        logger.info(f"{recipe_list}")
        # if not additive_list:
        #     return jsonify({'error': '缺少添加剂数据'}), 400

        # 生成配方
        try:
            logger.info(f"开始生成配方，recipe_list数量: {len(recipe_list)}, additive_list数量: {len(additive_list) if additive_list else 0}")
            formula_list = generate_initial_recipe(recipe_list, additive_list, electrolyte_quantity, additive_range)
            logger.info(f"配方生成完成，formula_list类型: {type(formula_list)}, 长度: {len(formula_list)}")
            if formula_list:
                logger.info(f"第一个配方结构: {formula_list[0]}")
                logger.info(f"第一个配方的keys: {formula_list[0].keys() if isinstance(formula_list[0], dict) else 'Not a dict'}")
                # 检查锂盐是否已正确添加
                if isinstance(formula_list[0], dict) and 'components' in formula_list[0]:
                    salt_comp = formula_list[0]['components'].get('salt_components', {})
                    logger.info(f"第一个配方的锂盐成分: {salt_comp}")
        except Exception as e:
            logger.error(f"生成配方失败: {str(e)}")
            logger.error(f"错误详情: ", exc_info=True)
            raise

        logger.info(f"最终formula_list: {formula_list}")

        # Li盐分开
        def reorganize_components(formula_list):
            """查找名字以 Li 开头的成分"""
            for formula in formula_list:
                comp = formula.get("components", {})
                recipe = comp.get("recipe_components", [])

                # 遍历查找名字以 Li 开头的
                for i, item in enumerate(recipe):
                    name = item.get("name", "")
                    if isinstance(name, str) and name.startswith("Li"):
                        # 找到就处理
                        li_item = recipe.pop(i)  # 删除并获取
                        comp["salt_components"] = [li_item]  # 移动到盐列表（改为列表形式）
                        break
                else:
                    # 没找到以 Li 开头的
                    comp["salt_components"] = []

            return formula_list
        formula_list = reorganize_components(formula_list)

        logger.info(f"分开后formula_list: {formula_list}")

        # 生成xlsx文件
        # 比例转质量
        quality_sum = quality + quality * additive_range
        data_list = formula_list

        for formula in data_list:
            components = formula.get("components", {})
            # 处理溶剂组分
            for component in components.get("recipe_components", []):
                M_quality = Chem.AddHs(Chem.MolFromSmiles(component["smiles"]))
                M_quality = round(Descriptors.MolWt(M_quality), 2)
                component["ratio"] = component["ratio"] * quality_sum / M_quality
            # 处理添加剂组分
            for component in components.get("additive_components", []):
                M_quality = Chem.AddHs(Chem.MolFromSmiles(component["smiles"]))
                M_quality = round(Descriptors.MolWt(M_quality), 2)
                component["ratio"] = component["ratio"] * quality_sum / M_quality
            # 处理锂盐组分（现在是列表格式）
            salt_components = components.get("salt_components", [])
            if salt_components and isinstance(salt_components, list):
                for salt in salt_components:
                    if isinstance(salt, dict) and "ratio" in salt:
                        M_quality = Chem.AddHs(Chem.MolFromSmiles(salt["smiles"]))
                        M_quality = round(Descriptors.MolWt(M_quality), 2)
                        salt["ratio"] = salt["ratio"] * quality_sum / M_quality

        for formula in formula_list:
            components = formula.get("components", {})
            # 处理溶剂组分
            for component in components.get("recipe_components", []):
                component["ratio"] = component["ratio"] * quality_sum
            # 处理添加剂组分
            for component in components.get("additive_components", []):
                component["ratio"] = component["ratio"] * quality_sum
            # 处理锂盐组分（现在是列表格式）
            salt_components = components.get("salt_components", [])
            if salt_components and isinstance(salt_components, list):
                for salt in salt_components:
                    if isinstance(salt, dict) and "ratio" in salt:
                        salt["ratio"] = salt["ratio"] * quality_sum

        logger.info(f"data_list:{data_list}")

        write_electrolyte_data_to_xlsx(data_list, r"D:\electrolyte\analysis\prediction\data\electrolyte_pre.xlsx")

        # 去运行prediction文件
        import subprocess
        old_dir = os.getcwd()
        folder_path ='D:\\electrolyte\\analysis\prediction'
        os.chdir(folder_path)
        subprocess.run(["python", r"D:\electrolyte\analysis\prediction\prediction.py"])
        os.chdir(old_dir)

        #从表格拿结果
        lce_array = extract_lce_from_excel(r"D:\electrolyte\analysis\prediction\electrolyte_pre_results.xlsx")

        logger.info(lce_array)
        for i in range(electrolyte_quantity):
            formula_list[i]["lce"] = lce_array[i]

        # for formula in formula_list:
        #     formula['components']['']
        return jsonify({
            'success': True,
            'formula': formula_list,
            "quality_sum": quality_sum,
            'message': '配方生成成功'
        })

    except Exception as e:
        logger.error(f"生成配方时出错: {str(e)}")
        return jsonify({'error': f'生成失败: {str(e)}'}), 500

@ai_designer_bp.route('/battery/start', methods=['POST'])
def battery_start1():
    """开始实验"""
    # -------------------------------------------------------
    # 1. 获取参数并校验
    # -------------------------------------------------------
    logger.info("开始实验")
    data = request.get_json()
    if not data:
        return jsonify({'error': '请求体为空'}), 400
        
    project_id = data.get("project_id")
    if not project_id:
        return jsonify({'error': '缺少 project_id'}), 400

    #配方数组
    formula_arr = data.get("formula_arr")
    if not formula_arr:
        return jsonify({'error': '缺少 formula_arr'}), 400

    #蓝电测试
    startLandian = data.get("startLandian")
    if not startLandian:
        return jsonify({'error': '缺少 startLandian'}), 400
    
    process_json_str = data.get("process_json_str")
    


    logger.info(f"准备启动项目：{project_id}")

    # -------------------------------------------------------
    # 2. 检查设备占用 (防御性检查)
    # -------------------------------------------------------
    try:
        # 确保 injector_obj_list 已经被正确导入或定义
        if not injector_obj_list:
            logger.warning(f"警告: injector_obj_list 为空，跳过占用检查")

        for tmp_injector in injector_obj_list:
            if not hasattr(tmp_injector, 'db_manager'): 
                continue
                
            tmp_order_dict = tmp_injector.db_manager.get_order(tmp_injector.project_id)
            current_status = tmp_order_dict.get("status")
            
            if current_status in ["启动", "组装中"]:
                logger.warning(f"设备忙碌中，当前状态: {current_status}")
                # 直接抛出异常，进入 except 处理
                raise Exception(f"当前有实验正在进行（{current_status}），请等待结束")
                
    except Exception as e:
        logger.error(f"无法启动设备 (检查阶段): {str(e)}")
        # 这里的 e 是有效的
        return jsonify({'error': f'无法启动: {str(e)}'}), 500

    # -------------------------------------------------------
    # 3. 真正启动设备 (执行阶段)
    # -------------------------------------------------------
    import sys
    import os

    # 1. 获取当前文件的绝对路径 (.../code/backend/api/...py)
    current_path = os.path.dirname(os.path.abspath(__file__))

    # 2. 第一次 dirname -> .../code/backend/api
    # 3. 第二次 dirname -> .../code/backend
    # 4. 第三次 dirname -> .../code (找到了项目的根目录！)
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_path)))

    # 5. 把项目根目录加入 Python 搜索路径
    sys.path.append(project_root)

    # 假设你的文件名叫 target_file.py，函数叫 my_func
    from fpxh_control_sdk.plc_control_new import equipment_start
    try:
        # 假设 equipment_start 返回 True 表示成功，False/报错 表示失败

        recipe_list = [
            {
                "bottle_no": int(key),            # 将字符串键 "1" 转为整数 1
                "channel_component": value        # value 就是对应的数组 [0, 0.0, ...]
            }
            for key, value in formula_arr.items() # 遍历字典的键值对
        ]

        arg_dict = {
            "bottle_mass_redundancy": 400,
            "order_num": project_id,
            "total_battery_count_each_bottle": 1,
            "battery_electrolyte_mass": 4200,
            "recipe_list": recipe_list
        }
        logger.info(f"开始实验传入参数{arg_dict}")
        
        result = equipment_start(arg_dict, startLandian, process_json_str=None)
        
        if result == True: # 注意这里是 True (大写)
            logger.info(f"项目 {project_id} 启动指令已发送")
    
            # # 启动任务前，先记录当前时间作为开始时间
            # start_time = datetime.now()
            # job_id = f"job_{project_id}"

            # # 添加定时任务：每5分钟执行一次
            # if not scheduler.get_job(job_id):
            #     scheduler.add_job(
            #         id=job_id,
            #         func=check_status_task,
            #         trigger='interval',
            #         minutes=0,
            #         args=[project_id, start_time, arg_dict],
            #         replace_existing=True
            #     )
            #     logger.info(f"实验已启动，后台监控中, order_num:{project_id}")
            # else:
            #     logger.info("该订单已在监控列表中")
            return jsonify({
                'success': True,
                'message': "成功开始实验"
            })
        else:
            # 如果函数返回了 False，说明启动失败，但没有报错
            # 这时候没有变量 e，我们需要自己写错误原因
            msg = "启动指令返回失败（未知原因）"
            logger.error(msg)
            return jsonify({'error': msg}), 500

    except Exception as e:
        # 只有这里的 except 才能捕获 equipment_start 里的崩溃
        logger.error(f"启动设备出错 (执行阶段崩溃): {str(e)}")
        return jsonify({'error': f'启动执行失败: {str(e)}'}), 500

# #定时任务
# def check_status_task(project_id, start_time, arg_dict, injector):
#     """定时检查逻辑"""
#     # 获取 job_id 方便后续删除
#     job_id = f"job_{project_id}"
#     logger.info("定时任务")

#     logger.info(f"injector:{injector}")

#     # 1. 检查是否超过一小时
#     if datetime.now() - start_time > timedelta(hours=1):
#         print(f"订单 {project_id} 检查超时（1小时），停止监控并报错")
#         if scheduler.get_job(job_id):
#             scheduler.remove_job(job_id)
#         return

#     # 2. 检查数据库状态
#     try:
#         order_dict = injector.db_manager.get_order(project_id)
#         status = order_dict.get("status")
#     except Exception as e:
#         print(f"查询订单 {project_id} 状态时发生异常: {str(e)}")
    
#     try:
#         if status == "组装完成": 
#             print(f"订单 {project_id} 状态 OK，开始测试")
#             injector.send_all_battery_test2landian()
            
#             # 任务完成，移除定时器
#             if scheduler.get_job(job_id):
#                 scheduler.remove_job(job_id)
#         else:
#             print(f"订单 {project_id} 当前状态: {status}，继续等待...")
            
#     except Exception as e:
#         print(f"开始订单 {project_id} 测试时发生异常: {str(e)}")

def smart_truncate(text, max_chars):
    """尽量在完整句子处截断"""
    if len(text) <= max_chars:
        return text
    
    # 尝试在句号处截断
    truncated = text[:max_chars]
    
    # 找最后一个合适的截断点
    truncate_points = ['。', '！', '？', '.', '!', '?', '\n\n', '\n']
    
    for point in truncate_points:
        last_index = truncated.rfind(point)
        if last_index > len(truncated) * 0.9:  # 在末尾附近找到
            return truncated[:last_index + len(point)]
    
    # 没找到，就在最后一个空格处截断
    last_space = truncated.rfind(' ')
    if last_space > len(truncated) * 0.9:
        return truncated[:last_space]
    
    return truncated

def write_electrolyte_data_to_xlsx(data_list, filename):
    """
    将电解质配方数据写入xlsx文件
    
    参数:
    data_list: 包含电解质配方数据的列表
    filename: 要写入的文件名
    """
    import openpyxl
    
    # 定义表头
    headers = [
        "Salt1_moles", "Salt2_moles", "Salt3_moles",
        "Solvent1_moles", "Solvent2_moles", "Solvent3_moles",
        "Additive1_moles", "Additive2_moles", "Additive3_moles",
        "Salt1_name", "Salt2_name", "Salt3_name",
        "Solvent1_name", "Solvent2_name", "Solvent3_name",
        "Additive1_name", "Additive2_name", "Additive3_name",
        "LCE"
    ]
    
    # 步骤1: 创建或清空工作簿，只保留表头
    try:
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Electrolyte Compositions"
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        print(f"已创建/清空文件 '{filename}'")
        
    except Exception as e:
        print(f"创建工作簿时出错: {e}")
        return
    
    # 步骤2: 解析数据并写入Excel
    try:
        # 从第2行开始写入数据（第1行是表头）
        current_row = 2
        
        # 遍历每个iteration的数据
        for item in data_list:
            iteration = item['iteration']
            components = item['components']
            
            # 初始化一行数据
            row_data = [""] * len(headers)
            
            # 处理盐组分 (salt_components)
            salt_components = components.get('salt_components', [])
            for i, salt in enumerate(salt_components[:3]):  # 最多取3个盐
                # 设置盐的摩尔量
                if i < 3:
                    row_data[i] = salt.get('ratio', '')  # Salt1_moles, Salt2_moles, Salt3_moles
                # 设置盐的名称
                row_data[9 + i] = salt.get('smiles', '')  # Salt1_name, Salt2_name, Salt3_name
            
            # 处理溶剂组分 (recipe_components)
            solvent_components = components.get('recipe_components', [])
            for i, solvent in enumerate(solvent_components[:3]):  # 最多取3个溶剂
                # 设置溶剂的摩尔量
                row_data[3 + i] = solvent.get('ratio', '')  # Solvent1_moles, Solvent2_moles, Solvent3_moles
                # 设置溶剂的名称
                row_data[12 + i] = solvent.get('smiles', '')  # Solvent1_name, Solvent2_name, Solvent3_name
            
            # 处理添加剂组分 (additive_components)
            additive_components = components.get('additive_components', [])
            for i, additive in enumerate(additive_components[:3]):  # 最多取3个添加剂
                # 设置添加剂的摩尔量
                row_data[6 + i] = additive.get('ratio', '')  # Additive1_moles, Additive2_moles, Additive3_moles
                # 设置添加剂的名称
                row_data[15 + i] = additive.get('smiles', '')  # Additive1_name, Additive2_name, Additive3_name
            
            # LCE字段暂时留空
            row_data[17] = ""  # LCE (注意：索引从0开始，所以17对应第18列)
            
            # 写入数据到Excel
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_idx, value=value)
            
            current_row += 1
        
        # 保存工作簿
        wb.save(filename)
        print(f"成功写入 {len(data_list)} 行数据到文件 '{filename}'")
        
    except Exception as e:
        print(f"写入数据时出错: {e}")
        return
    
def get_csv_responsen(csv_file):
    """"cvs转数组"""
    # 读取 CSV
    # header=None: 告诉 pandas 这个文件没有表头
    # names=['ID', 'SMILES']: 手动指定列名，这样生成的 JSON key 就是 "ID" 和 "SMILES"
    # sep=',': 指定分隔符为逗号
    import pandas as pd
    df = pd.read_csv(csv_file, sep=',')
    # 3. 处理空值 (防止 JSON 报错)
    df = df.where(pd.notnull(df), None)

    result = df.to_dict(orient='records')

    return result

def generate_initial_recipe(recipe_source, additive_source, electrolyte_quantity, additive_range=None):
    """
    生成初始电解质配方

    参数:
    recipe_source: 配方源成分列表（可能包含溶剂和锂盐）
    additive_source: 添加剂源成分列表
    electrolyte_quantity: 生成的配方数量
    additive_range: 添加剂总占比（单个数值），默认为 0.01-0.05 之间的随机值
    """
    import random
    import logging
    logger = logging.getLogger(__name__)

    logger.info(f"开始生成配方，recipe_source长度: {len(recipe_source) if recipe_source else 0}")
    logger.info(f"recipe_source内容: {recipe_source}")

    # additive_range 现在是单个数值
    if additive_range is None:
        additive_value = random.uniform(0.01, 0.05)  # 默认随机值
    else:
        additive_value = float(additive_range)  # 直接使用传入的数值

    # recipe_source中的所有成分（包括锂盐）都当作溶剂处理
    solvent_source = []

    if recipe_source:
        for item in recipe_source:
            if isinstance(item, dict):
                name = item.get('name', '')
                # 检查ratio是否为空或0，如果是则随机生成
                ratio = item.get('ratio')
                if ratio is None or ratio == 0:
                    item['ratio'] = random.uniform(0.5, 2.0)
                    logger.info(f"{name} 的ratio为空或0，已随机生成: {item['ratio']}")

                solvent_source.append(item)
                logger.info(f"添加成分: {name}")
            else:
                logger.warning(f"recipe_source中的项不是字典: {type(item)}")

    logger.info(f"recipe_source总数量: {len(solvent_source)}")

    # 全随机模式函数（可复用）
    def distribute_random_weights(items, target_percent):
        """全随机分配整个target_percent"""
        if not items:
            return []
        
        # 生成随机权重
        weights = [random.uniform(0.5, 2.0) for _ in range(len(items))]
        total_weight = sum(weights)
        
        # 归一化到目标比例
        result = []
        for i, item in enumerate(items):
            name = item.get('name', f'Component_{i}')
            smiles = item.get('smiles', '')
            if total_weight > 0:
                final_ratio = (weights[i] / total_weight) * target_percent
            else:
                final_ratio = target_percent / len(items)
            result.append({'name': name, 'ratio': round(final_ratio, 6), 'smiles': smiles})
        
        return result
    
    # 全固定比例模式函数
    def distribute_all_fixed_weights(items, target_percent):
        """全固定比例分配"""
        if not items:
            return []
            
        result = []
        fixed_total = 0
        
        # 计算所有固定比例的总和
        for item in items:
            ratio = item.get('ratio')
            if ratio is not None and isinstance(ratio, (int, float)):
                fixed_total += float(ratio)
        
        # 按原始比例分配target_percent
        for item in items:
            name = item.get('name', 'Unknown')
            smiles = item.get('smiles', '')
            ratio = float(item.get('ratio', 0))
            if fixed_total > 0:
                final_ratio = (ratio / fixed_total) * target_percent
            else:
                final_ratio = target_percent / len(items)
            result.append({'name': name, 'ratio': round(final_ratio, 6), 'smiles': smiles})
        
        return result
    
    # 混合模式函数（有ratio的固定，没有ratio的随机）
    def distribute_mixed_weights(items, target_percent):
        """混合模式：有ratio的按比例分配，没有ratio的随机分配"""
        if not items:
            return []
            
        # 分离固定比例和随机比例的成分
        fixed_items = []
        random_items = []
        fixed_total = 0
        
        for item in items:
            name = item.get('name', 'Unknown')
            smiles = item.get('smiles', '')
            ratio = item.get('ratio')
            
            if ratio is not None and isinstance(ratio, (int, float)):
                # 固定比例成分
                fixed_ratio = float(ratio)
                fixed_total += fixed_ratio
                fixed_items.append({
                    'name': name, 
                    'original_ratio': fixed_ratio,
                    'smiles': smiles
                })
            else:
                # 随机比例成分
                random_items.append({'name': name, 'smiles': smiles})
        
        result = []
        
        # 情况1：有固定比例成分，也有随机成分（混合模式）
        if fixed_items and random_items:
            # 固定比例部分先分配（占target_percent的70%-90%）
            fixed_portion = random.uniform(0.7, 0.9)  # 固定部分比例
            random_portion = 1 - fixed_portion  # 随机部分比例
            
            # 分配固定比例部分
            for item in fixed_items:
                if fixed_total > 0:
                    # 按原始比例关系分配固定部分
                    fixed_share = item['original_ratio'] / fixed_total
                    final_ratio = fixed_share * (target_percent * fixed_portion)
                else:
                    final_ratio = (target_percent * fixed_portion) / len(fixed_items)
                result.append({'name': item['name'], 'ratio': round(final_ratio, 6), 'smiles': item.get('smiles', '')})
            
            # 分配随机比例部分
            if random_items:
                weights = [random.uniform(0.5, 2.0) for _ in range(len(random_items))]
                total_weight = sum(weights)
                
                for i, item in enumerate(random_items):
                    if total_weight > 0:
                        final_ratio = (weights[i] / total_weight) * (target_percent * random_portion)
                    else:
                        final_ratio = (target_percent * random_portion) / len(random_items)
                    result.append({'name': item['name'], 'ratio': round(final_ratio, 6), 'smiles': item.get('smiles', '')})
        
        # 情况2：全部是固定比例
        elif fixed_items and not random_items:
            # 按原始比例关系分配整个target_percent
            for item in fixed_items:
                if fixed_total > 0:
                    final_ratio = (item['original_ratio'] / fixed_total) * target_percent
                else:
                    final_ratio = target_percent / len(fixed_items)
                result.append({'name': item['name'], 'ratio': round(final_ratio, 6), 'smiles': item.get('smiles', '')})
        
        # 情况3：全部是随机比例
        elif not fixed_items and random_items:
            # 全随机分配整个target_percent
            random_result = distribute_random_weights([{'name': item['name'], 'smiles': item['smiles']} for item in random_items], target_percent)
            result.extend(random_result)
        
        return result
    
    # 检查是否所有成分都有固定比例
    def all_have_fixed_ratios(items):
        """检查列表中所有成分都有固定比例"""
        if not items:
            return False
            
        for item in items:
            ratio = item.get('ratio')
            if ratio is None or not isinstance(ratio, (int, float)):
                return False
        return True
    
    # 检查是否有混合成分（部分有ratio，部分没有）
    def has_mixed_ratios(items):
        """检查是否有混合情况（部分有ratio，部分没有）"""
        if not items:
            return False
            
        has_fixed = False
        has_random = False
        
        for item in items:
            ratio = item.get('ratio')
            if ratio is not None and isinstance(ratio, (int, float)):
                has_fixed = True
            else:
                has_random = True
            
            if has_fixed and has_random:
                return True
        return False
    
    # 检查是否有随机成分（没有ratio）
    def has_random_ratios(items):
        """检查是否有随机成分"""
        if not items:
            return False
            
        for item in items:
            ratio = item.get('ratio')
            if ratio is None or not isinstance(ratio, (int, float)):
                return True
        return False
    
    # 判断recipe和additive的模式
    recipe_all_fixed = all_have_fixed_ratios(recipe_source) if recipe_source else False
    additive_all_fixed = all_have_fixed_ratios(additive_source) if additive_source else False
    all_have_fixed = recipe_all_fixed and additive_all_fixed
    
    recipe_mixed = has_mixed_ratios(recipe_source) if recipe_source else False
    additive_mixed = has_mixed_ratios(additive_source) if additive_source else False
    has_mixed = recipe_mixed or additive_mixed
    
    recipe_has_random = has_random_ratios(recipe_source) if recipe_source else False
    additive_has_random = has_random_ratios(additive_source) if additive_source else False
    has_any_random = recipe_has_random or additive_has_random
    
    # 确定整体模式
    if all_have_fixed:
        # 模式1：所有成分都有固定比例
        mode = "all_fixed"
    elif has_mixed:
        # 模式2：混合模式（部分固定，部分随机）
        mode = "mixed"
    else:
        # 模式3：所有成分都随机
        mode = "all_random"

    # 验证additive_value参数
    if additive_value < 0 or additive_value > 1:
        raise ValueError("添加剂比例必须在 0 到 1 之间")

    # 返回结果
    final_result = []

    # 开始 electrolyte_quantity 次循环
    for i in range(electrolyte_quantity):
        # 检查 additive_source 是否为空
        if not additive_source:
            # additive_source为空，solvent占全部权重（100%）
            solvent_target = 1.0  # 100%
            additive_target = 0.0  # 0%
            # additive_result 应为空列表
            additive_result = []
        else:
            # additive_source不为空，使用固定数值
            additive_target = additive_value
            solvent_target = 1 - additive_target

        # 判断当前迭代使用什么模式（基于溶剂和添加剂）
        solvent_all_fixed = all_have_fixed_ratios(solvent_source) if solvent_source else False
        additive_all_fixed = all_have_fixed_ratios(additive_source) if additive_source else False
        all_have_fixed = solvent_all_fixed and additive_all_fixed

        solvent_mixed = has_mixed_ratios(solvent_source) if solvent_source else False
        additive_mixed = has_mixed_ratios(additive_source) if additive_source else False
        has_mixed = solvent_mixed or additive_mixed

        if all_have_fixed:
            current_mode = "all_fixed"
        elif has_mixed:
            current_mode = "mixed"
        else:
            current_mode = "all_random"

        # 根据模式分配溶剂和添加剂
        if current_mode == "all_fixed":
            if i == 0:
                solvent_result = distribute_all_fixed_weights(solvent_source, solvent_target)
                if additive_source:
                    additive_result = distribute_all_fixed_weights(additive_source, additive_target)
                else:
                    additive_result = []
            else:
                solvent_result = distribute_random_weights(solvent_source, solvent_target)
                if additive_source:
                    additive_result = distribute_random_weights(additive_source, additive_target)
                else:
                    additive_result = []

        elif current_mode == "mixed":
            solvent_result = distribute_mixed_weights(solvent_source, solvent_target)
            if additive_source:
                additive_result = distribute_mixed_weights(additive_source, additive_target)
            else:
                additive_result = []

        else:  # all_random
            solvent_result = distribute_random_weights(solvent_source, solvent_target)
            if additive_source:
                additive_result = distribute_random_weights(additive_source, additive_target)
            else:
                additive_result = []

        # 创建返回结果字典，直接添加锂盐
        final_result.append({
            "iteration": i + 1,
            "lce": None,
            "components": {
                "recipe_components": solvent_result,
                "additive_components": additive_result,
                "salt_components": {}
            }
        })

    logger.info(f"配方生成完成，共生成{len(final_result)}个配方")

    return final_result

def _extract_keywords_from_parameters(data):
    """
    从电池参数数据中提取关键词
    
    Args:
        data: 包含电池参数的字典列表
        
    Returns:
        list: 提取的关键词列表，格式为自然语言描述
    """
    keywords = []
    
    # 中英文映射字典
    material_map = {
        'NMC811': 'NCM811',
        'NMC523': 'NCM523',
        'LFP': 'LFP',
        'LCO': 'LCO',
        '石墨': 'graphite',
        'SiC': 'SiC',
        'Li': 'lithium'
    }
    
    application_map = {
        '动力': 'electric vehicles',
        '蓄能': 'energy storage',
        '3C': '3C products'
    }
    
    safety_map = {
        1: 'low safety',
        2: 'basic safety',
        3: 'standard safety',
        4: 'high safety',
        5: 'very high safety'
    }
    
    for item in data:
        category = item.get('category', '')
        param_name = item.get('param_name', '')
        param_value = item.get('param_value', '')
        unit = item.get('unit', '')
        
        # 提取基本信息类关键词
        if category == '基本信息':
            if param_name == '正极材料' and param_value:
                eng_material = material_map.get(param_value, param_value)
                keywords.append(f"{eng_material} cathode material")
            
            elif param_name == '负极材料' and param_value:
                eng_material = material_map.get(param_value, param_value)
                keywords.append(f"{eng_material} anode material")
            
            elif param_name == '应用场景' and param_value:
                eng_application = application_map.get(param_value, param_value)
                if eng_application == 'electric vehicles':
                    keywords.append(eng_application)
                    keywords.append("high safety requirements")
                elif eng_application == 'energy storage':
                    keywords.append("energy storage battery")
        
        # 提取性能参数类关键词
        elif category == '性能参数':
            if param_name == '循环寿命' and param_value:
                if unit:
                    keywords.append(f"cycle life {param_value} {unit}")
                else:
                    keywords.append(f"cycle life {param_value} cycles")
            
            elif param_name == '能量密度' and param_value:
                if unit:
                    keywords.append(f"energy density {param_value} {unit}")
                else:
                    keywords.append(f"energy density {param_value} Wh/kg")
            
            elif param_name == '功率密度' and param_value:
                if unit:
                    keywords.append(f"power density {param_value} {unit}")
            
            elif param_name == '安全性' and param_value:
                safety_level = safety_map.get(param_value, f"safety level {param_value}")
                if param_value >= 4:
                    keywords.append("high safety requirements")
                keywords.append(f"{safety_level}")
            
            elif param_name == '工作温度' and param_value:
                # 提取温度范围信息
                temp_range = item.get('range', '')
                if temp_range:
                    keywords.append(f"operating temperature range {temp_range}")
    
    # 根据应用场景添加相关关键词
    for item in data:
        if item.get('category') == '基本信息' and item.get('param_name') == '应用场景':
            app_value = item.get('param_value', '')
            if app_value == '蓄能':
                keywords.append("long lifespan low cost")
    
    # 去重并保持原有顺序
    unique_keywords = []
    for kw in keywords:
        if kw not in unique_keywords:
            unique_keywords.append(kw)
    
    return unique_keywords

def _call_bm25_search(keywords):
    """调用BM25文献搜索算法"""
    try:
        # 导入BM25搜索函数
        import sys
        sys.path.append(r'D:\electrolyte\BO')

        # 检查文件是否存在
        import os
        bm25_file = r'D:\electrolyte\BO\dbsearch.py'
        if not os.path.exists(bm25_file):
            raise Exception(f"BM25搜索文件不存在: {bm25_file}")

        logger.info(f"BM25搜索文件存在: {bm25_file}")

        # 动态导入BM25搜索模块
        try:
            newsearch_file = r'D:\electrolyte\BO\dbsearch.py'
            if not os.path.exists(newsearch_file):
                raise Exception(f"BM25搜索文件不存在: {newsearch_file}")

            spec = importlib.util.spec_from_file_location("newsearch", newsearch_file)
            newsearch_module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(newsearch_module)
            bm25_search_from_postgres = newsearch_module.bm25_search_from_postgres
            logger.info("BM25搜索函数动态导入成功")
        except Exception as e:
            logger.error(f"导入BM25搜索函数失败: {str(e)}")
            raise Exception(f"无法导入BM25搜索函数: {str(e)}")

        # 数据库配置
        db_config = {
            'dbname': 'postgres',
            'user': 'postgres',
            'password': '123456',
            'host': 'localhost',
            'port': 5432
        }

        logger.info(f"开始调用BM25搜索，关键词: {keywords}")
        logger.info(f"数据库配置: {db_config}")

        # 调用BM25搜索（注意参数顺序要与函数定义一致）
        results, titles = bm25_search_from_postgres(
            dbname=db_config['dbname'],
            user=db_config['user'],
            password=db_config['password'],
            host=db_config['host'],
            port=db_config['port'],
            keywords=keywords,
            top_k=10
        )

        logger.info(f"BM25搜索返回 {len(results)} 个结果, {len(titles)} 个标题")

        # 格式化结果为前端需要的格式
        formatted_results = []
        for i, result in enumerate(results):
            # 过滤无期刊的文献
            journal = result.get('journal', '')
            if not journal or journal.strip() == '':
                logger.info(f"跳过无期刊信息的文献: {result.get('title', 'Unknown')}")
                continue

            formatted_result = {
                'id': f"bm25_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{i}",
                'title': result.get('title', ''),
                'authors': result.get('authors', ''),
                'journal': journal,
                'publication_year': result.get('year', 2023),
                'doi': result.get('doi', ''),
                'abstract': result.get('abstract', ''),
                'relevance_score': float(result.get('score', 0.0)),
                'processing_status': 'completed',
                'extraction_confidence': min(1.0, float(result.get('score', 0.0)) / 10.0),  # 归一化置信度
                'keywords': result.get('keyword', ''),  # 从数据库中获取关键词字段
                'extracted_formulas': [],  # BM25主要做文献匹配，公式提取留给后续步骤
                'extracted_molecules': [],
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat()
            }
            formatted_results.append(formatted_result)

        logger.info(f"BM25文献匹配完成，找到 {len(formatted_results)} 篇相关文献")

        # 添加文献数据库统计信息
        literature_stats = {
            'total_literature_count': len(results),  # 查询到的文献总数
            'processed_count': len(formatted_results),  # 实际处理的文献数（过滤无期刊后的数量）
            'filtered_count': len(results) - len(formatted_results),  # 被过滤的无期刊文献数
            'query_timestamp': datetime.now().isoformat(),
            'keywords_used': keywords
        }

        # 将统计信息添加到每个结果中
        for result in formatted_results:
            result['literature_stats'] = literature_stats

        return formatted_results

    except ImportError as e:
        logger.error(f"无法导入BM25搜索模块: {str(e)}")
        raise Exception(f"BM25搜索模块不可用: {str(e)}")
    except Exception as e:
        error_str = str(e)
        logger.error(f"BM25搜索调用失败: {error_str}")

        # 检查具体错误类型并给出更友好的错误信息
        if "connection" in error_str.lower():
            logger.error("数据库连接失败，请检查PostgreSQL是否运行")
            raise Exception("数据库连接失败，请检查数据库服务是否运行")
        elif "password" in error_str.lower() or "authentication" in error_str.lower():
            logger.error("数据库认证失败，请检查用户名和密码")
            raise Exception("数据库认证失败，请检查数据库配置")
        elif "database" in error_str.lower() and "not exist" in error_str.lower():
            logger.error("数据库不存在，请检查数据库名称")
            raise Exception("数据库不存在，请检查数据库配置")
        elif "no such table" in error_str.lower():
            logger.error("papers表不存在，请先创建数据库表")
            raise Exception("数据库表不存在，请检查数据库初始化")
        elif "syntax error" in error_str.lower():
            logger.error("SQL语法错误")
            raise Exception("数据库查询语法错误")
        else:
            logger.error(f"未知错误: {error_str}")
            raise Exception(f"BM25搜索失败: {error_str}")

def _build_user_request_description(confirmed_parameters):
    """根据确认的参数构建用户请求描述（优化版，兼容列表和字典格式）"""
    description_parts = []

    # 如果是列表类型，从参数记录中提取信息
    if isinstance(confirmed_parameters, list):
        # 从参数列表中提取关键信息
        system_value = ''
        app_value = ''
        perf_reqs = {}

        for param in confirmed_parameters:
            if isinstance(param, dict):
                category = param.get('category', '')
                param_name = param.get('param_name', '')
                param_value = param.get('param_value', '')

                # 提取体系类型
                if '体系' in category or 'system' in category.lower():
                    system_value = param_value

                # 提取应用场景
                elif '应用' in category or 'application' in category.lower():
                    app_value = param_value

                # 提取性能参数
                elif '性能' in category or 'performance' in category.lower():
                    perf_reqs[param_name] = param_value

    # 兼容前端的简化格式：{system_type, application_scenario, performance_requirements}
    # 和原来的详细格式：{basic_info: {system_type, application_scenario}, performance_params}
    elif isinstance(confirmed_parameters, dict):
        if 'system_type' in confirmed_parameters:
            # 前端简化格式
            system_value = confirmed_parameters.get('system_type', '')
            app_value = confirmed_parameters.get('application_scenario', '')
            perf_reqs = confirmed_parameters.get('performance_requirements', {})
        else:
            # 原来的详细格式
            basic_info = confirmed_parameters.get('basic_info', {})
            system_type_obj = basic_info.get('system_type', {})
            system_value = system_type_obj.get('value', '') if isinstance(system_type_obj, dict) else system_type_obj

            application_obj = basic_info.get('application_scenario', {})
            app_value = application_obj.get('value', '') if isinstance(application_obj, dict) else application_obj

            perf_reqs = confirmed_parameters.get('performance_params', {})
            # 转换 performance_params 格式到统一的 performance_requirements 格式
            temp_perf_reqs = {}
            for param_key, param_data in perf_reqs.items():
                if isinstance(param_data, dict):
                    temp_perf_reqs[param_key] = param_data.get('value')
                else:
                    temp_perf_reqs[param_key] = param_data
            perf_reqs = temp_perf_reqs
    else:
        # 不支持的格式
        logger.error(f"不支持的参数类型: {type(confirmed_parameters)}")
        return "寻找高性能电池材料和分子结构"

    # 体系类型描述
    if system_value:
        description_parts.append(f"目标电池体系：{system_value}")

    # 应用场景描述
    if app_value:
        # 应用场景映射（从英文value转为中文显示）
        app_mapping = {
            'electric_vehicle': '电动汽车',
            'energy_storage': '储能系统',
            'consumer_electronics': '消费电子',
            'aerospace': '航空航天'
        }
        # 如果app_value已经是中文，直接使用；否则映射
        chinese_app = app_mapping.get(app_value, app_value)
        description_parts.append(f"应用场景：{chinese_app}")

    # 性能参数描述（更详细）
    performance_desc = []

    if perf_reqs and isinstance(perf_reqs, dict):
        for param_key, value in perf_reqs.items():
            if value is not None:
                # 性能参数中文标签映射
                param_label_mapping = {
                    'energy_density': '目标能量密度',
                    'target_energy_density': '目标能量密度',
                    'power_density': '功率密度',
                    'cycle_life': '循环寿命',
                    'working_temperature': '工作温度',
                    'safety': '安全性'
                }
                chinese_label = param_label_mapping.get(param_key, param_key)

                # 获取单位
                unit = ''
                if param_key == 'energy_density' or param_key == 'target_energy_density':
                    unit = 'Wh/kg'
                elif param_key == 'power_density':
                    unit = 'W/kg'
                elif param_key == 'cycle_life':
                    unit = '次'
                elif param_key == 'working_temperature':
                    unit = '°C'

                chinese_desc = f"{chinese_label}：{value}{unit}"
                performance_desc.append(chinese_desc)

    if performance_desc:
        description_parts.append("关键性能要求：" + "，".join(performance_desc))

    # 安全性要求和成本约束（仅当confirmed_parameters是字典类型时处理）
    if isinstance(confirmed_parameters, dict):
        # 安全性要求
        safety_requirements = confirmed_parameters.get('safety_requirements', {})
        if isinstance(safety_requirements, dict):
            safety_desc = []
            for req_key, req_data in safety_requirements.items():
                if isinstance(req_data, dict) and req_data.get('value'):
                    req_label = req_data.get('label', req_key)
                    safety_desc.append(req_label)

            if safety_desc:
                description_parts.append("安全性要求：" + "，".join(safety_desc))

        # 成本约束
        cost_constraints = confirmed_parameters.get('cost_constraints', {})
        if isinstance(cost_constraints, dict):
            target_cost = cost_constraints.get('target_cost', {})
            if isinstance(target_cost, dict) and target_cost.get('value'):
                cost_label = target_cost.get('label', '')
                if cost_label:
                    description_parts.append(f"成本约束：{cost_label}")

    # 添加具体搜索目标
    search_goals = [
        "寻找符合上述要求的电池材料、电解液配方和分子结构",
        "重点关注高性能、高安全性的电解液的添加剂",
        "筛选具有良好电化学性能的分子化合物",
        "不挖盐，salts直接返回空就行",
        "不挖溶剂，solvents也直接返回空",
        "控制一下挖掘的分子个数，只挖掘比较优秀的2个分子，不要水、含Li的，要求去重",
        '''additives要求返回格式如下：{
            "Ethylene carbonate": "C1COC(=O)O1",
            "Dimethyl carbonate": "COC(=O)OC",
            "Diethyl carbonate": "CCOC(=O)OCC"
        }'''
    ]
    description_parts.extend(search_goals)

    # 如果没有有效描述，使用默认描述
    if not description_parts:
        return "寻找高性能电池材料和分子结构，重点关注电解液配方"

    # 组合完整描述
    full_description = "。".join(description_parts)

    logger.info(f"构建的用户请求描述: {full_description}")

    return full_description

def quick_lookup(smiles):
    """快速查询CAS、名称、简称"""
    try:
        # 标准化SMILES
        mol = Chem.MolFromSmiles(smiles)
        if mol:
            smiles = Chem.MolToSmiles(mol, canonical=True)

        # 查询PubChem
        compounds = pcp.get_compounds(smiles, 'smiles')
        if not compounds:
            return None

        compound = compounds[0]

        # 提取CAS号
        cas_numbers = []
        common_names = []

        # 安全获取synonyms列表
        synonyms_list = []
        if hasattr(compound, 'synonyms') and compound.synonyms is not None:
            if isinstance(compound.synonyms, (list, tuple)):
                synonyms_list = compound.synonyms
            else:
                logger.warning(f"synonyms 不是列表类型: {type(compound.synonyms)}")

        for syn in synonyms_list:
            # CAS号通常为 xxx-xx-x 格式
            if '-' in syn and syn.replace('-', '').isdigit():
                parts = syn.split('-')
                if len(parts) == 3:
                    cas_numbers.append(syn)
            # 常用名通常是较短的英文名称
            elif len(syn.split()) <= 3 and syn.replace(' ', '').isalpha():
                common_names.append(syn)

        # 获取IUPAC名称
        iupac_name = getattr(compound, 'iupac_name', None)

        return {
            'SMILES': smiles,
            'CAS': cas_numbers[:3],  # 最多3个CAS
            'IUPAC': iupac_name,
            'common_name': common_names[0] if common_names else None,
            'synonyms': synonyms_list[:5],  # 最多返回5个同义词
            'formula': getattr(compound, 'molecular_formula', None),
            'weight': getattr(compound, 'molecular_weight', None)
        }

    except Exception as e:
        print(f"查询失败: {e}")
        return None
    
def extract_lce_from_excel(file_path, sheet_name=0, lce_column='LCE'):
    """
    从Excel文件中提取LCE列数据
    
    参数:
    file_path: Excel文件路径
    sheet_name: 工作表名称或索引，默认为第一个工作表
    lce_column: LCE列的名称，默认为'LCE'
    
    返回:
    LCE列数据的数组
    """
    try:
        # 读取Excel文件
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        # 检查LCE列是否存在
        if lce_column not in df.columns:
            # 如果没有找到LCE列，尝试查找最后一列
            print(f"警告: 未找到名为 '{lce_column}' 的列")
            print(f"可用的列: {list(df.columns)}")
            
            # 尝试获取最后一列
            last_column = df.columns[-1]
            print(f"使用最后一列: '{last_column}'")
            lce_column = last_column
        
        # 提取LCE列数据，转换为列表
        lce_data = df[lce_column].dropna().tolist()
        
        return lce_data
    
    except FileNotFoundError:
        print(f"错误: 找不到文件 '{file_path}'")
        return []
    except Exception as e:
        print(f"读取文件时出错: {e}")
        return []